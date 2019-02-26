from django.shortcuts import render, redirect
from django.shortcuts import render_to_response
from django.http import HttpResponse
from django.template import loader,Context

from django.db.models import Avg, Sum, Max, Min, Count
from django.db.models import F, Q
# from django.db.models import Sum,Count

import requests, json, time, datetime, hashlib, random
import csv, codecs
# from data.test import *

import os
import re
from requests.cookies import RequestsCookieJar

from PIL import Image

from django.conf import settings

from web.models import base_conf, base_user_sign_log
from web.models import pureftp
from web.models import hr_department, hr_hr, employee_department
from web.models import asset_conf, asset_category, asset_property, position, asset_attachment, asset_application, asset_allot

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from openpyxl import load_workbook

def partt(id):
    #这是一个定义函数，返回父级关系
    tid = asset_category.objects.filter(id=id).first()
    if tid.parentid:
        disname = partt(tid.parentid.id) + ' / ' +  tid.name
        return disname
    else:
        return tid.name

def status(id,show):
    if show == 0:
        statusname=['', '闲置', '在用', '维修', '报废']
    elif show == 1:
        statusname=['',
                    '<span class="text-secondary"><i class="fa fa-circle-o" aria-hidden="true"></i></span>',
                    '<span class="text-success"><i class="fa fa-check-circle" aria-hidden="true"></i></span>',
                    '<span class="text-warning"><i class="fa fa-exclamation-triangle" aria-hidden="true"></i></span>',
                    '<span class="text-danger"><i class="fa fa-minus-circle" aria-hidden="true"></i></span>']
    elif show == 2:
        statusname=['',
                    '<span class="badge badge-secondary">闲置</span>',
                    '<span class="badge badge-success">在用</span>',
                    '<span class="badge badge-warning">维修</span>',
                    '<span class="badge badge-danger">报废</span>']


    elif show == 10:
        statusname=['', '审批中', '已通过', '已驳回', '已取消', '', '通过后撤销']
    elif show == 11:
        statusname=['',
                    '<span class="text-success"><i class="fa fa-code-fork" aria-hidden="true"></i></span>',
                    '<span class="text-success"><i class="fa fa-check-circle" aria-hidden="true"></i></span>',
                    '<span class="text-danger"><i class="fa fa-minus-circle" aria-hidden="true"></i></span>',
                    '<span class="text-danger"><i class="fa fa-times-circle" aria-hidden="true"></i></span>',
                    '<span class="text-light"><i class="fa fa-times-circle" aria-hidden="true"></i></span>',
                    '<span class="text-warning"><i class="fa fa-exclamation-triangle" aria-hidden="true"></i></span>']
    elif show == 12:
        statusname = ['',
                      '<span class="badge badge-success">审批中</span>',
                      '<span class="badge badge-secondary">已通过</span>',
                      '<span class="badge badge-danger">已驳回</span>',
                      '<span class="badge badge-light">已取消</span>',
                      '<span class="badge badge-light"> </span>',
                      '<span class="badge badge-warning">通过后撤销</span>',
                      ]
    return statusname[id]

# Create your views here.

def index(request):
    context={}

    username = request.COOKIES.get('usercookie', None)
    if username:
        try:
            signuser = hr_hr.objects.get(session=username)
        except Exception:
            context['userinfo'] = '用户'
            return render(request, 'sign.html', context)
        context['userinfo'] = signuser.name
    else:
        context['userinfo'] = '用户'
        return render(request, 'sign.html', context)

    return render(request, 'base.html', context)

def asset_applicant_list(request):
    request.encoding = 'utf-8'
    context={}
    context['title']='申请单'

    ps = asset_application.objects.filter(active=True).values('appltno',
                                                              'appdate',
                                                              'applicant',
                                                              'applicant__name',
                                                              'applicant__employee_department__departmentid__name',
                                                              'Explain',
                                                              'needasset',
                                                              'type',
                                                              'backdate',
                                                              'status',
                                                              'flow',
                                                              'active').order_by('status','-appdate')
    print(ps.count())

    s = []
    for t in list(ps):
        if t['type']:
            if t['type'] == 1:
                # t['type'] = '<span class="badge badge-success">领用</span>'
                t['type'] = '<span class="text-success">领用</span>'
            if t['type'] == 2:
                # t['type'] = '<span class="badge badge-warning">借用</span>'
                t['type'] = '<span class="text-warning">借用</span>'
        if t['status']:
            if t['status']:
                t['statusstr'] = status(t['status'], 11)
        s.append(t)

    context['context'] = s
    # print(ps)

    return render(request, 'asset_applicant_list.html', context)

def asset_allot_list(request):
    # 领用列表模块
    request.encoding = 'utf-8'
    context={}
    context['title']='领用单'

    ps = asset_allot.objects.all().values()

    context['context'] = ps
    # print(ps)

    return render(request, 'asset_allot_list.html', context)

def asset_config(request):
    context={}

    username = request.COOKIES.get('usercookie', None)
    if username:
        try:
            signuser = hr_hr.objects.get(session=username)
        except Exception:
            context['userinfo'] = '用户'
            return render(request, 'sign.html', context)
        context['userinfo'] = signuser.name
    else:
        context['userinfo'] = '用户'
        return render(request, 'sign.html', context)

    return render(request, 'asset_conf.html', context)

def asset_kanban_board(request):
    context={}

    username = request.COOKIES.get('usercookie', None)
    if username:
        try:
            signuser = hr_hr.objects.get(session=username)
        except Exception:
            context['userinfo'] = '用户'
            return render(request, 'sign.html', context)
        context['userinfo'] = signuser.name
    else:
        context['userinfo'] = '用户'
        return render(request, 'sign.html', context)

    if request.method == "GET":
        print(request.GET)
        if "act" in request.GET:
            if request.GET['act'] == 'pi':
                f = asset_property.objects.filter(bom=False,active=True).values('status').annotate(count=Count('status')).order_by()[:10]
                for item in f:
                    item['disname'] = status(item['status'],0)
                if f :
                    data = {}
                    data['code'] = 0
                    data['msg'] = "ok"
                    data['data'] = list(f)
                    data = json.dumps(data)
                else:
                    data={}
                    data['code'] = 1
                    data['msg'] = "Fail"
                    data['data'] = "无返回值"
                    data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == 'y_num':
                f = asset_property.objects.extra(select={'year': 'strftime("%%Y",purchase)'}).values('year').annotate(count=Count('purchase')).order_by('-year')
                if f :
                    data = {}
                    data['code'] = 0
                    data['msg'] = "ok"
                    data['data'] = list(f)
                    data = json.dumps(data)
                else:
                    data={}
                    data['code'] = 1
                    data['msg'] = "Fail"
                    data['data'] = "无返回值"
                    data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == 'y_price':
                f = asset_property.objects.extra(select={'year': 'strftime("%%Y",purchase)'}).values('year').annotate(price=Sum('price')).order_by('-year')
                if f :
                    data = {}
                    data['code'] = 0
                    data['msg'] = "ok"
                    data['data'] = list(f)
                    data = json.dumps(data)
                else:
                    data={}
                    data['code'] = 1
                    data['msg'] = "Fail"
                    data['data'] = "无返回值"
                    data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == 'dep_num':
                f = hr_department.objects.all().values('name').annotate(num=Count('employee_department__employeeid__asset_property__sid'),picre=Sum('employee_department__employeeid__asset_property__price')).order_by('-num')[:10]
                if f :
                    data = {}
                    data['code'] = 0
                    data['msg'] = "ok"
                    data['data'] = list(f)
                    data = json.dumps(data)
                else:
                    data={}
                    data['code'] = 1
                    data['msg'] = "Fail"
                    data['data'] = "无返回值"
                    data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")

    return render(request, 'asset_kanban_board.html', context)

def asset_property_list(request):
    request.encoding = 'utf-8'
    context={}
    context['title']='设备列表'

    username = request.COOKIES.get('usercookie', None)
    if username:
        try:
            signuser = hr_hr.objects.get(session=username)
        except Exception:
            context['userinfo'] = '用户'
            return render(request, 'sign.html', context)
        context['userinfo'] = signuser.name
    else:
        context['userinfo'] = '用户'
        return render(request, 'sign.html', context)

    if request.method == "GET":
        print('GET')
        print(request.GET)
        if "act" in request.GET:
            if request.GET['act'] == 'filter':
                data = {}
                search = {}
                search['field'] = request.GET['field']
                search['ilike'] = request.GET['ilike']
                if search['field'] == "active":
                    search['ilike'] = int(search['ilike'])
                    ps = asset_property.objects.filter(active=search['ilike']).values('id',
                                                                            'status',
                                                                            'sid',
                                                                            'name',
                                                                            'specifications',
                                                                            'purchase',
                                                                            'warranty',
                                                                            'user__name',
                                                                            'user__active',
                                                                            'position',
                                                                            'sn').order_by('name')
                    u = list(ps)
                elif search['field'] != "all":
                    if(search['ilike'] == 'null'):
                        search['ilike'] = None
                    kwargs={}
                    kwargs['active'] = True
                    # kwargs['bom'] = False
                    kwargs[search['field']]=search['ilike']
                    ps = asset_property.objects.filter(**kwargs).values('id',
                                                                        'status',
                                                                        'sid',
                                                                        'name',
                                                                        'specifications',
                                                                        'purchase',
                                                                        'warranty',
                                                                        'user__name',
                                                                        'user__active',
                                                                        'position',
                                                                        'sn').order_by('name')
                    u = list(ps)
                else:
                    ps = asset_property.objects.filter(active=True,bom=False).filter(Q(user__name__icontains=search['ilike'])
                                                                                     | Q(sid__icontains=search['ilike'])
                                                                                     | Q(sn__icontains=search['ilike'])
                                                                                     | Q(name__icontains=search['ilike'])
                                                                                     | Q(specifications__icontains=search['ilike'])
                                                                                     | Q(position__icontains=search['ilike'])).values('id',
                                                                                                                                      'status',
                                                                                                                                      'sid',
                                                                                                                                      'name',
                                                                                                                                      'specifications',
                                                                                                                                      'purchase',
                                                                                                                                      'warranty',
                                                                                                                                      'user__name',
                                                                                                                                      'user__active',
                                                                                                                                      'position',
                                                                                                                                      'sn')
                    print(ps.count())
                    u = list(ps)

                s = []
                for t in u:
                    if t['purchase']:
                        t['purchase'] = t['purchase'].strftime("%Y-%m-%d")
                    if t['warranty']:
                        t['warranty'] = t['warranty'].strftime("%Y-%m-%d")
                    if t['status']:
                        t['statusstr'] = status(t['status'] ,2)
                    s.append(t)
                if len(s):
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['spk'] = len(s)
                    data['data'] = s
                else:
                    data['code'] = 1
                    data['msg'] = "Fail"
                    data['spk'] = 0
                    data['data'] = "无返回值"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            # elif request.GET['act'] == 'sort':
            #     data = {}
            #     sns = request.GET['field']
            #     ps = asset_property.objects.filter(bom=False,active=True).values('id',
            #                                                            'status',
            #                                                            'sid',
            #                                                            'name',
            #                                                            'specifications',
            #                                                            'purchase',
            #                                                            'warranty',
            #                                                            'user__name',
            #                                                            'user__active',
            #                                                            'position',
            #                                                            'user__employee_department__departmentid__name',
            #                                                            'sn').order_by(sns)
            #     s = []
            #     u = list(ps)
            #     for t in u:
            #         if t['purchase']:
            #             t['purchase'] = t['purchase'].strftime("%Y-%m-%d")
            #         if t['warranty']:
            #             t['warranty'] = t['warranty'].strftime("%Y-%m-%d")
            #         s.append(t)
            #
            #     if len(s):
            #         data['code'] = 0
            #         data['msg'] = "OK"
            #         data['spk'] = len(s)
            #         data['data'] = s
            #     else:
            #         data['code'] = 1
            #         data['msg'] = "Fail"
            #         data['spk'] = 0
            #         data['data'] = "无返回值"
            #
            #     data = json.dumps(data)
            #     return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == 'groupby':
                data = {}
                groupby = request.GET['field']
                ps = asset_property.objects.filter(bom=False,active=True).values(groupby).annotate(number=Count('id'),pice=Sum('price')).order_by(groupby)

                s = []
                u = list(ps)
                for t in u:
                    t['name'] = t[groupby]
                    t['disn'] = t[groupby]
                    if t[groupby] == '':
                        t['name'] = "未定义"
                        t['disn'] = "未定义"
                    t['val'] = t[groupby]

                    t['field'] = groupby
                    if t['field'] == "categoryid":
                        if t[groupby]:
                            t['disn'] = partt(t[groupby])
                        else:
                            t['disn'] = "未定义"
                    if t['field'] == "status":
                        t['disn'] = status(t[groupby],0)
                    s.append(t)

                if len(s):
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['spk'] = len(s)
                    data['data'] = s
                else:
                    data['code'] = 1
                    data['msg'] = "Fail"
                    data['spk'] = 0
                    data['data'] = "无返回值"

                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == 'enableparts':
                data = {}
                search = {}
                search['field'] = request.GET['field']
                search['ilike'] = request.GET['ilike']
                # if search['field'] == "active":
                search['ilike'] = int(search['ilike'])
                pn = asset_property.objects.filter(active=True, categoryid=search['ilike'], status=1).values('id',
                                                                                  'status',
                                                                                  'sid',
                                                                                  'name',
                                                                                  'specifications',
                                                                                  'purchase',
                                                                                  'warranty',
                                                                                  'user__name',
                                                                                  'user__active',
                                                                                  'position',
                                                                                  'sn').order_by('name')
                s = []
                for t in list(pn):
                    if t['purchase']:
                        t['purchase'] = t['purchase'].strftime("%Y-%m-%d")
                    if t['warranty']:
                        t['warranty'] = t['warranty'].strftime("%Y-%m-%d")
                    if t['status']:
                        t['statusstr'] = status(t['status'], 2)
                    s.append(t)
                if len(s):
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['spk'] = len(s)
                    data['data'] = s
                else:
                    data['code'] = 1
                    data['msg'] = "Fail"
                    data['spk'] = 0
                    data['data'] = "无返回值"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
        else:
            # 没有动作,输出默认列表
            ps = asset_property.objects.filter(bom=False, active=True).filter(Q(asset_attachment__final=True)
                                                                              | Q(asset_attachment__final=None)).values('id',
                                                                                                                        'status',
                                                                                                                        'sid',
                                                                                                                        'name',
                                                                                                                        'specifications',
                                                                                                                        'sn',
                                                                                                                        'warranty',
                                                                                                                        'user__name',
                                                                                                                        'user__active',
                                                                                                                        'asset_attachment__filepath',
                                                                                                                        'asset_attachment__thumbnail',
                                                                                                                        'asset_attachment__final',
                                                                                                                        'position',
                                                                                                                        'user__employee_department__departmentid__name')
            context['spk'] = ps.count()

            # 排序
            orderby = request.GET.get('s', 'name')
            context['sort'] = orderby
            ps = ps.order_by(orderby)

            # 页码
            page = request.GET.get('p', 1)
            context['page'] = page

            #每页条目数
            baseconfig = asset_conf.objects.get(pk=1)
            lpnum = request.GET.get('n', baseconfig.boardnum)
            context['lpnum'] = lpnum

            #显示方式 board、list
            typeviewlist = ["list","board","details"]
            tview = int(request.GET.get('v', baseconfig.viewtype))

            context['tview'] = tview
            context['tviewstr'] = typeviewlist[tview-1]

            paginator = Paginator(ps, lpnum)

            try:
                data = paginator.page(page)
            except PageNotAnInteger:
                data = paginator.page(1)
            except EmptyPage:
                data = paginator.page(paginator.num_pages)
            ps = data

            # print(data.next_page_number())
            context['fpk'] = data.start_index()
            context['tpk'] = data.end_index()

            try:
                context['pagprevious'] = data.previous_page_number()
            except EmptyPage:
                context['pagprevious'] = 0

            try:
                context['pagnext'] = data.next_page_number()
            except EmptyPage:
                context['pagnext'] = 0

            s = []
            for t in list(ps):
                # if t['warranty']:
                #     t['warranty'] = t['warranty'].strftime("%Y-%m-%d")
                if t['status']:
                    t['statusstr'] = status(t['status'], 2)
                if t['asset_attachment__filepath'] == None:
                    t['asset_attachment__filepath'] = '/static/img/asset.png'
                if t['asset_attachment__thumbnail'] == None:
                    t['asset_attachment__thumbnail'] = '/static/img/asset.png'
                s.append(t)
            context['context'] = s

            # context['context'] = list(ps)

            # print(context)
            # print(context['context'])
            return render(request, 'asset_property_list.html', context)
    elif request.method == "POST":
        print("POST")
        print(request.POST)

        # awg=request.scheme + '://127.0.0.1:8000' + request.path + '?'
        # for i in request.POST:
        #     awg=awg + i + '=' + request.POST[i] +'&'
        # return HttpResponse(awg[:-1])
    return render(request, 'asset_property_list.html', context)

def asset_property_sub_board(request):
    request.encoding = 'utf-8'
    context={}
    context['title']='设备列表'

    username = request.COOKIES.get('usercookie', None)
    if username:
        try:
            signuser = hr_hr.objects.get(session=username)
        except Exception:
            context['userinfo'] = '用户'
            return render(request, 'sign.html', context)
        context['userinfo'] = signuser.name
    else:
        context['userinfo'] = '用户'
        return render(request, 'sign.html', context)

    print(request.GET)
    print(request.POST)
    ps = asset_property.objects.filter(bom=False,active=True).filter(Q(asset_attachment__final=True)
                                                                     |Q(asset_attachment__final=None)).values('id',
                                                                                                              'status',
                                                                                                              'sid',
                                                                                                              'name',
                                                                                                              'specifications',
                                                                                                              'warranty',
                                                                                                              'user__name',
                                                                                                              'user__active',
                                                                                                              'asset_attachment__filepath',
                                                                                                              'asset_attachment__thumbnail',
                                                                                                              'asset_attachment__final',
                                                                                                              'position',
                                                                                                              'user__employee_department__departmentid__name',
                                                                                                              'sn').order_by('name', 'specifications', 'sid')

    context['spk'] = ps.count()

    # 每页条目数
    baseconfig = asset_conf.objects.get(pk=1)
    lpnum = baseconfig.boardnum

    page = request.POST.get('p', 1)

    paginator = Paginator(ps,lpnum)
    data=paginator.page(page)

    s = []
    for t in list(data):
        # if t['warranty']:
        #     t['warranty'] = t['warranty'].strftime("%Y-%m-%d")
        if t['status']:
            t['statusstr'] = status(t['status'], 2)
        if t['asset_attachment__filepath'] == None:
            t['asset_attachment__filepath'] = '/static/img/asset.png'
        if t['asset_attachment__thumbnail'] == None:
            t['asset_attachment__thumbnail'] = '/static/img/asset.png'
        s.append(t)
    context['context'] = s

    # 显示方式 board、list
    typeviewlist = ["list", "board", "details"]

    context['tview'] = tview = typeviewlist[int(request.POST.get('v', 1)) - 1]

    view_tpl = 'asset_property_sub_' + tview + '.html'

    # t = loader.get_template(view_tpl)
    # m = t.render(context)
    # print(m)

    # data = json.dumps(data)
    # return HttpResponse(data, content_type="application/json")

    return render(request, view_tpl, context)

def asset_property_form(request):
    request.encoding = 'utf-8'
    context={}
    context['title']='设备单'

    username = request.COOKIES.get('usercookie', None)
    if username:
        try:
            signuser = hr_hr.objects.get(session=username)
        except Exception:
            context['userinfo'] = '用户'
            return render(request, 'sign.html', context)
        context['userinfo'] = signuser.name
    else:
        context['userinfo'] = '用户'
        return render(request, 'sign.html', context)

    if request.method == "GET":
        print(request.GET)
        if "act" in request.GET:
            if request.GET['act'] == "display":
                context['act'] = "display"

                # 当前设备id、前后id
                pn = int(request.GET['id'])
                ppn = pn-1
                npn = pn+1
                context['pk'] = pn
                context['ppk'] = ppn
                context['npk'] = npn

                #设备总数
                spn = asset_property.objects.filter(active=True).count()
                context['spk'] = spn

                # 设备类型DisplayName计算
                cats = asset_category.objects.filter(active=True).values('id', 'name', 'displayname').order_by('displayname')
                context['cats'] = cats

                # 可用人员查询
                hrs = hr_hr.objects.filter(active=True).order_by('name')
                context['hrs']= hrs

                try:
                    # 当前设备信息
                    ps = asset_property.objects.get(id=pn)
                except Exception:
                    return redirect('/property_list/')
                else:
                    context['context'] = ps

                    # 设备配件获取
                    prs = ps.asset_property_set.all()
                    context['parts'] = prs
                    context['partsnum'] = prs.count()

                # 设备照片信息
                asspic = asset_attachment.objects.filter(property=pn,active=True, final=True).first()
                if asspic:
                    context['imgid'] = asspic.id
                    context['headimg'] = asspic.filepath
                else:
                    # 如果无图则传占位符
                    context['imgid'] = 0
                    context['headimg'] = 'holder.js/64x64'

                # 单据标题
                context['title'] = '设备单'
            elif request.GET['act'] == "disheadimg":
                pn = int(request.GET['id'])
                ps = asset_attachment.objects.filter(property=pn, active=True).values("id","name","filepath")

                msg = {}
                if ps.count():
                    # 返回JSON数据格式 0:为正常
                    # {"code": 0, "msg": "OK", "data": [1,2,3,]}
                    msg['code'] = 0
                    msg['msg'] = "OK"

                    # 附加数据
                    msg['data'] = list(ps)
                else:
                    msg['code'] = 1
                    msg['msg'] = "Fiale"
                    msg['data'] = None

                data = json.dumps(msg)

                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "dishrinfo":
                ghrid = int(request.GET['hrid'])
                hrdeppos = employee_department.objects.filter(employeeid__id=ghrid).values('employeeid__id','employeeid__name','employeeid__position','departmentid__name').first()
                data = {}
                if hrdeppos:
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['data'] = hrdeppos
                else:
                    data['code'] = 1
                    data['msg'] = "FAIL"
                    data['data'] = "None"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "discategory":
                gcategoryid = int(request.GET['categoryid'])
                categorydeppos = asset_category.objects.filter(id=gcategoryid).values('id', 'name', 'bom').first()
                data = {}
                if categorydeppos:
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['data'] = categorydeppos
                else:
                    data['code'] = 1
                    data['msg'] = "FAIL"
                    data['data'] = "None"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "indexpartscategoryid":
                # parts = asset_property.objects.filter(categoryid__active=True, categoryid__bom=True,active=True).values('categoryid__id','categoryid__name').annotate(Count('id'))
                parts = asset_property.objects.filter(active=True,categoryid__active=True, categoryid__bom=True, status=1).values('categoryid__id','categoryid__name').annotate(count=Count('id'))
                # pp = asset_category.objects.filter(bom=True,active=True,asset_property__active=True).values("id", "name","asset_property__active").annotate(Count('asset_property__id'))
                # print(pp)

                # 改名，校正数据格式
                s = []
                for t in list(parts):
                    t['disn'] = partt(t['categoryid__id'])
                    t['id'] = t['categoryid__id']
                    t['num'] = t['count']
                    s.append(t)


                data = {}
                if parts:
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['data'] = s
                else:
                    data['code'] = 1
                    data['msg'] = "FAIL"
                    data['data'] = "None"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "indexname":
                val = request.GET['ilike']
                t = asset_property.objects.filter(name__icontains=val,active=True).values('name').distinct()
                data = {}
                if t:
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['data'] = list(t)
                else:
                    data['code'] = 1
                    data['msg'] = "Fail"
                    data['data'] = "无返回值"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "indexmodel":
                val = request.GET['ilike']
                t = asset_property.objects.filter(model__icontains=val,active=True).values('model').distinct()
                data = {}
                if t:
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['data'] = list(t)
                else:
                    data['code'] = 1
                    data['msg'] = "Fail"
                    data['data'] = "无返回值"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "indexspec":
                val = request.GET['ilike']
                t = asset_property.objects.filter(specifications__icontains=val,active=True).values('specifications').distinct()
                data = {}
                if t:
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['data'] = list(t)
                else:
                    data['code'] = 1
                    data['msg'] = "Fail"
                    data['data'] = "无返回值"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "indexposition":
                val = request.GET['ilike']
                t = asset_property.objects.filter(position__icontains=val,active=True).values('position').distinct().order_by('position')
                data = {}
                if t:
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['data'] = list(t)
                else:
                    data['code'] = 1
                    data['msg'] = "Fail"
                    data['data'] = "无返回值"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "chacksid":
                val = request.GET['sid']
                cid = request.GET['id']
                t = asset_property.objects.filter(sid=val).exclude(pk=cid).first()
                data = {}
                if t:
                    data['code'] = 1
                    data['msg'] = "fail"
                    data['data'] = "输入值已存在"
                else:
                    data['code'] = 0
                    data['msg'] = "OK"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "chacksn":
                val = request.GET['sn']
                cid = request.GET['id']
                t = asset_property.objects.filter(sn=val).exclude(pk=cid).first()
                data = {}
                if t:
                    data['code'] = 1
                    data['msg'] = "fail"
                    data['data'] = "输入值已存在"
                else:
                    data['code'] = 0
                    data['msg'] = "OK"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "create":
                context['act'] = "create"
                context['pk'] = id = int(request.GET['id'])

                # 新建设置默认值
                #
                cats = asset_category.objects.filter(active=True).values('id', 'name', 'displayname').order_by('displayname')
                context['cats'] = cats

                hrs = hr_hr.objects.filter(active=True).order_by('name')
                context['hrs'] = hrs

                ps = {}
                # 默认编号
                # Todo 这编号在建立规则后，可以根据规则生成
                # ps['sid'] = 0


                # Todo  id = 0 为新建 其它为复制
                # ======================================
                if id:
                    # 取现有值进行复制
                    context['title'] = '设备单 / 复制'
                    ps = asset_property.objects.filter(id=id).first()
                else:
                    context['title'] = '设备单 / 新建'

                    # 默认类型
                    ps['categoryid'] = 1

                    # 默认价格
                    ps["price"] = 0

                    # 默认出厂、维保、报废日期
                    t =datetime.datetime.now()
                    ps["manufacture"] = ps["purchase"] = ps["warranty"] = t

                context['context'] = ps
                return render(request, 'asset_property_form.html', context)
            elif request.GET['act'] == "edit":
                pass
    elif request.method == "POST":
        print(request.POST)
        if "act" in request.POST:
            if request.POST['act'] == "active":
                assid = int(request.POST['id'])
                act = asset_property.objects.filter(id=assid).update(active=True)
                return HttpResponse(act)
            if request.POST['act'] == 'unactive':
                assid = int(request.POST['id'])
                act = asset_property.objects.filter(id=assid).update(active=False)
                return HttpResponse(act)
            if request.POST['act'] == 'delimg':
                id = int(request.POST['id'])
                a = asset_attachment.objects.filter(id=id).update(active=False,final=False)
                pid = int(request.POST['pid'])

                n = asset_attachment.objects.filter(property=pid,active=True).order_by('-id').first()
                data = {}
                # 返回1.失败1 2.成功：无图2、有图0
                if n:
                    n.final = True
                    n.save()
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['id'] = n.id
                    data['filepath'] = n.filepath
                else:
                    data['code'] = 2
                    data['msg'] = "OK"
                    data['id'] = 0

                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            if request.POST['act'] == 'create':
                act='create'

                ugcatid = int(request.POST['categoryid'])
                if ugcatid == 0:
                    catid = None
                else:
                    catid = asset_category.objects.filter(id=ugcatid).first()

                guserid = int(request.POST['user'])
                if guserid == 0:
                    userid = None
                else:
                    userid = hr_hr.objects.filter(id=guserid).first()

                if (request.POST.get('bom', False)):
                    bom = False
                else:
                    bom = True

                # 获取默认日期，当天日期
                pdate = request.POST['purchase']
                if pdate == '':
                    purchase = datetime.datetime.today()
                else:
                    purchase = datetime.datetime.strptime(pdate, "%Y-%m-%d")

                wdat = request.POST['warranty']
                if wdat == '':
                    warranty = datetime.datetime.today()
                else:
                    warranty = datetime.datetime.strptime(wdat, "%Y-%m-%d")

                rdat = request.POST['manufacture']
                if rdat == '':
                    manufacture = datetime.datetime.today()
                else:
                    manufacture = datetime.datetime.strptime(rdat, "%Y-%m-%d")

                item = asset_property(
                    sid = request.POST['sid'],
                    name = request.POST['name'],
                    sn=request.POST['sn'],
                    specifications = request.POST['specifications'],
                    model = request.POST['model'],
                    categoryid = catid,
                    bom = bom,
                    purchase = purchase,
                    price = float(request.POST['price']),
                    manufacture = manufacture,
                    warranty = warranty,
                    user = userid,
                    # partlist = request.POST[''],
                    position = request.POST['position'],
                    status = 1,
                    nots = request.POST['comment'],
                    active = True
                )

                item.save()
                id = item.id
                return HttpResponse(id)
            if request.POST['act'] == 'edit':
                act='edit'
                id = int(request.POST['id'])

                pid = asset_property.objects.get(id=id)

                ugcatid = int(request.POST['categoryid'])
                pid.categoryid = asset_category.objects.filter(id=ugcatid).first()

                guserid = int(request.POST['user'])
                pid.user = hr_hr.objects.filter(id=guserid).first()

                if (request.POST.get('bom', False)):
                    pid.bom = True
                else:
                    pid.bom = False

                print(pid.bom)

                # # 获取默认日期，当天日期
                pdate = request.POST['purchase']
                if pdate == '':
                    pid.purchase = datetime.datetime.today()
                else:
                    pid.purchase = datetime.datetime.strptime(pdate, "%Y-%m-%d")

                wdat = request.POST['warranty']
                if wdat == '':
                    pid.warranty = datetime.datetime.today()
                else:
                    pid.warranty = datetime.datetime.strptime(wdat, "%Y-%m-%d")

                rdat = request.POST['manufacture']
                if rdat == '':
                    pid.manufacture = datetime.datetime.today()
                else:
                    pid.manufacture = datetime.datetime.strptime(rdat, "%Y-%m-%d")

                pid.sid = request.POST.get('sid')
                pid.name = request.POST['name']
                pid.sn = request.POST['sn']
                pid.specifications = request.POST['specifications']
                pid.model = request.POST['model']
                pid.price = float(request.POST['price'])
                # # partlist = request.POST[''],
                pid.position = request.POST['position']
                # pid.status = 1
                pid.nots = request.POST['comment']
                # pid.active = True

                pid.save()

                return HttpResponse(id)
            if request.POST['act'] == 'addparts':
                assetidlist= request.POST['id']
                parentid = int(request.POST['parentid'])

                # Todo 向设备添加配件，1.改配件的父级 2.此配件归档 3.修改配件状态为使用中 4. 刷新对应设备的配件列表
                pps= asset_property.objects.get(id=parentid)

                for assetid in assetidlist.split(",")[0:-1]:
                    print(assetid)
                    ps = asset_property.objects.get(id=int(assetid))

                    ps.parentid = pps
                    # ps.active = False
                    ps.status = 2

                    ps.save()

                data={}

                data['code'] = 0
                data['msg'] = "ok"

                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            if request.POST['act'] == 'delparts':
                assetid = int(request.POST['id'])

                ps = asset_property.objects.get(id=assetid)
                ps.parentid = None
                # ps.active = True
                ps.status = 1
                ps.save()

                data = {}
                data['code'] = 0
                data['msg'] = "ok"

                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
    return render(request, 'asset_property_form.html', context)

def asset_parts_list(request):
    context={}
    context['title']='配件列表'

    username = request.COOKIES.get('usercookie', None)
    if username:
        try:
            signuser = hr_hr.objects.get(session=username)
        except Exception:
            context['userinfo'] = '用户'
            return render(request, 'sign.html', context)
        context['userinfo'] = signuser.name
    else:
        context['userinfo'] = '用户'
        return render(request, 'sign.html', context)

    ps = asset_property.objects.filter(bom=True,active=True).values().order_by('name', 'specifications', 'status')

    u = list(ps)
    s = []
    for t in u:
        if t['purchase']:
            t['purchase'] = t['purchase'].strftime("%Y-%m-%d")
        if t['warranty']:
            t['warranty'] = t['warranty'].strftime("%Y-%m-%d")
        if t['status']:
            t['statusstr'] = status(t['status'], 1)
        s.append(t)
        # print(s)
    context['context'] = s

    return render(request, 'parts_list.html', context)

def asset_parts_form(request):
    request.encoding = 'utf-8'
    context={}
    context['title']='配件单'

    username = request.COOKIES.get('usercookie', None)
    if username:
        try:
            signuser = hr_hr.objects.get(session=username)
        except Exception:
            context['userinfo'] = '用户'
            return render(request, 'sign.html', context)
        context['userinfo'] = signuser.name
    else:
        context['userinfo'] = '用户'
        return render(request, 'sign.html', context)

    if request.method == "GET":
        print(request.GET)
        if "act" in request.GET:
            if request.GET['act'] == "display":
                context['act'] = "display"

                # 当前设备id、前后id
                pn = int(request.GET['id'])
                ppn = pn-1
                npn = pn+1
                context['pk'] = pn
                context['ppk'] = ppn
                context['npk'] = npn

                #设备总数
                spn = asset_property.objects.filter(active=True).count()
                context['spk'] = spn

                # 设备类型DisplayName计算
                cats = asset_category.objects.filter(active=True).order_by('parentid','name')
                lcats = []
                for cat in cats:
                    ncat = {}
                    ncat["id"] = cat.id
                    ncat["name"] = cat.name
                    ncat["displayname"] =partt(cat.id)
                    lcats.append(ncat)
                context['cats'] = lcats

                # 可用人员查询
                hrs = hr_hr.objects.filter(active=True).order_by('name')
                context['hrs']= hrs

                try:
                    # 当前设备信息
                    ps = asset_property.objects.get(id=pn)
                except Exception:
                    return redirect('/property_list/')
                else:
                    context['context'] = ps

                    # 设备配件获取
                    prs = ps.asset_property_set.all()
                    context['parts'] = prs
                    context['partsnum'] = prs.count()

                # 设备照片信息
                asspic = asset_attachment.objects.filter(property=pn,active=True, final=True).first()
                if asspic:
                    context['imgid'] = asspic.id
                    context['headimg'] = asspic.filepath
                else:
                    # 如果无图则传占位符
                    context['imgid'] = 0
                    context['headimg'] = 'holder.js/64x64'

                # 单据标题
                context['title'] = '设备单'
            elif request.GET['act'] == "disheadimg":
                pn = int(request.GET['id'])
                ps = asset_attachment.objects.filter(property=pn, active=True).values("id","name","filepath")

                # 返回JSON数据格式 0:为正常
                # {"code": 0, "msg": "OK", "data": [1,2,3,]}
                msg = {}
                msg['code'] = 0
                msg['msg'] = "OK"

                # 附加数据
                ldata = list(ps)
                msg['data'] = ldata
                data = json.dumps(msg)

                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "dishrinfo":
                ghrid = int(request.GET['hrid'])
                hrdeppos = employee_department.objects.filter(employeeid__id=ghrid).values('employeeid__id','employeeid__name','employeeid__position','departmentid__name').first()
                data = {}
                if hrdeppos:
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['data'] = hrdeppos
                else:
                    data['code'] = 1
                    data['msg'] = "FAIL"
                    data['data'] = "None"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "discategory":
                gcategoryid = int(request.GET['categoryid'])
                categorydeppos = asset_category.objects.filter(id=gcategoryid).values('id', 'name', 'bom').first()
                data = {}
                if categorydeppos:
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['data'] = categorydeppos
                else:
                    data['code'] = 1
                    data['msg'] = "FAIL"
                    data['data'] = "None"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "indexpartscategoryid":
                # parts = asset_property.objects.filter(categoryid__active=True, categoryid__bom=True,active=True).values('categoryid__id','categoryid__name').annotate(Count('id'))
                parts = asset_property.objects.filter(active=True,categoryid__active=True, categoryid__bom=True, status=1).values('categoryid__id','categoryid__name').annotate(count=Count('id'))
                # pp = asset_category.objects.filter(bom=True,active=True,asset_property__active=True).values("id", "name","asset_property__active").annotate(Count('asset_property__id'))
                # print(pp)

                # 改名，校正数据格式
                s = []
                for t in list(parts):
                    t['disn'] = partt(t['categoryid__id'])
                    t['id'] = t['categoryid__id']
                    t['num'] = t['count']
                    s.append(t)


                data = {}
                if parts:
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['data'] = s
                else:
                    data['code'] = 1
                    data['msg'] = "FAIL"
                    data['data'] = "None"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "indexname":
                val = request.GET['ilike']
                t = asset_property.objects.filter(name__icontains=val,active=True).values('name').distinct()
                data = {}
                if t:
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['data'] = list(t)
                else:
                    data['code'] = 1
                    data['msg'] = "Fail"
                    data['data'] = "无返回值"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "indexmodel":
                val = request.GET['ilike']
                t = asset_property.objects.filter(model__icontains=val,active=True).values('model').distinct()
                data = {}
                if t:
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['data'] = list(t)
                else:
                    data['code'] = 1
                    data['msg'] = "Fail"
                    data['data'] = "无返回值"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "indexspec":
                val = request.GET['ilike']
                t = asset_property.objects.filter(specifications__icontains=val,active=True).values('specifications').distinct()
                data = {}
                if t:
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['data'] = list(t)
                else:
                    data['code'] = 1
                    data['msg'] = "Fail"
                    data['data'] = "无返回值"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "indexposition":
                val = request.GET['ilike']
                t = asset_property.objects.filter(position__icontains=val,active=True).values('position').distinct().order_by('position')
                data = {}
                if t:
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['data'] = list(t)
                else:
                    data['code'] = 1
                    data['msg'] = "Fail"
                    data['data'] = "无返回值"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "chacksid":
                val = request.GET['sid']
                cid = request.GET['id']
                t = asset_property.objects.filter(sid=val).exclude(pk=cid).first()
                data = {}
                if t:
                    data['code'] = 1
                    data['msg'] = "fail"
                    data['data'] = "输入值已存在"
                else:
                    data['code'] = 0
                    data['msg'] = "OK"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "chacksn":
                val = request.GET['sn']
                cid = request.GET['id']
                t = asset_property.objects.filter(sn=val).exclude(pk=cid).first()
                data = {}
                if t:
                    data['code'] = 1
                    data['msg'] = "fail"
                    data['data'] = "输入值已存在"
                else:
                    data['code'] = 0
                    data['msg'] = "OK"
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            elif request.GET['act'] == "create":
                context['act'] = "create"
                context['pk'] = id = int(request.GET['id'])

                # 新建设置默认值
                #
                cats = asset_category.objects.filter(active=True).order_by('parentid', 'name')
                lcats = []
                for cat in cats:
                    ncat = {}
                    ncat["id"] = cat.id
                    ncat["name"] = cat.name
                    ncat["displayname"] = partt(cat.id)
                    lcats.append(ncat)
                context['cats'] = lcats

                hrs = hr_hr.objects.filter(active=True).order_by('name')
                context['hrs'] = hrs

                ps = {}
                # 默认编号
                # Todo 这编号在建立规则后，可以根据规则生成
                # ps['sid'] = 0


                # Todo  id = 0 为新建 其它为复制
                # ======================================
                if id:
                    # 取现有值进行复制
                    context['title'] = '设备单 / 复制'
                    ps = asset_property.objects.filter(id=id).first()
                else:
                    context['title'] = '设备单 / 新建'

                    # 默认类型
                    ps['categoryid'] = 1

                    # 默认价格
                    ps["price"] = 0

                    # 默认出厂、维保、报废日期
                    t =datetime.datetime.now()
                    ps["manufacture"] = ps["purchase"] = ps["warranty"] = t

                context['context'] = ps
                return render(request, 'asset_property_form.html', context)
            elif request.GET['act'] == "edit":
                pass
    elif request.method == "POST":
        print(request.POST)
        if "act" in request.POST:
            if request.POST['act'] == "active":
                assid = int(request.POST['id'])
                act = asset_property.objects.filter(id=assid).update(active=True)
                return HttpResponse(act)
            if request.POST['act'] == 'unactive':
                assid = int(request.POST['id'])
                act = asset_property.objects.filter(id=assid).update(active=False)
                return HttpResponse(act)
            if request.POST['act'] == 'delimg':
                id = int(request.POST['id'])
                a = asset_attachment.objects.filter(id=id).update(active=False,final=False)
                pid = int(request.POST['pid'])

                n = asset_attachment.objects.filter(property=pid,active=True).order_by('-id').first()
                data = {}
                # 返回1.失败1 2.成功：无图2、有图0
                if n:
                    n.final = True
                    n.save()
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['id'] = n.id
                    data['filepath'] = n.filepath
                else:
                    data['code'] = 2
                    data['msg'] = "OK"
                    data['id'] = 0

                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            if request.POST['act'] == 'create':
                act='create'

                ugcatid = int(request.POST['categoryid'])
                if ugcatid == 0:
                    catid = None
                else:
                    catid = asset_category.objects.filter(id=ugcatid).first()

                guserid = int(request.POST['user'])
                if guserid == 0:
                    userid = None
                else:
                    userid = hr_hr.objects.filter(id=guserid).first()

                if (request.POST.get('bom', False)):
                    bom = False
                else:
                    bom = True

                # 获取默认日期，当天日期
                pdate = request.POST['purchase']
                if pdate == '':
                    purchase = datetime.datetime.today()
                else:
                    purchase = datetime.datetime.strptime(pdate, "%Y-%m-%d")

                wdat = request.POST['warranty']
                if wdat == '':
                    warranty = datetime.datetime.today()
                else:
                    warranty = datetime.datetime.strptime(wdat, "%Y-%m-%d")

                rdat = request.POST['manufacture']
                if rdat == '':
                    manufacture = datetime.datetime.today()
                else:
                    manufacture = datetime.datetime.strptime(rdat, "%Y-%m-%d")

                item = asset_property(
                    sid = request.POST['sid'],
                    name = request.POST['name'],
                    sn=request.POST['sn'],
                    specifications = request.POST['specifications'],
                    model = request.POST['model'],
                    categoryid = catid,
                    bom = bom,
                    purchase = purchase,
                    price = float(request.POST['price']),
                    manufacture = manufacture,
                    warranty = warranty,
                    user = userid,
                    # partlist = request.POST[''],
                    position = request.POST['position'],
                    status = 1,
                    nots = request.POST['comment'],
                    active = True
                )

                item.save()
                id = item.id
                return HttpResponse(id)
            if request.POST['act'] == 'edit':
                act='edit'
                id = int(request.POST['id'])

                pid = asset_property.objects.get(id=id)

                ugcatid = int(request.POST['categoryid'])
                pid.categoryid = asset_category.objects.filter(id=ugcatid).first()

                guserid = int(request.POST['user'])
                pid.user = hr_hr.objects.filter(id=guserid).first()

                if (request.POST.get('bom', False)):
                    pid.bom = True
                else:
                    pid.bom = False

                print(pid.bom)

                # # 获取默认日期，当天日期
                pdate = request.POST['purchase']
                if pdate == '':
                    pid.purchase = datetime.datetime.today()
                else:
                    pid.purchase = datetime.datetime.strptime(pdate, "%Y-%m-%d")

                wdat = request.POST['warranty']
                if wdat == '':
                    pid.warranty = datetime.datetime.today()
                else:
                    pid.warranty = datetime.datetime.strptime(wdat, "%Y-%m-%d")

                rdat = request.POST['manufacture']
                if rdat == '':
                    pid.manufacture = datetime.datetime.today()
                else:
                    pid.manufacture = datetime.datetime.strptime(rdat, "%Y-%m-%d")

                pid.sid = request.POST.get('sid')
                pid.name = request.POST['name']
                pid.sn = request.POST['sn']
                pid.specifications = request.POST['specifications']
                pid.model = request.POST['model']
                pid.price = float(request.POST['price'])
                # # partlist = request.POST[''],
                pid.position = request.POST['position']
                # pid.status = 1
                pid.nots = request.POST['comment']
                # pid.active = True

                pid.save()

                return HttpResponse(id)
            if request.POST['act'] == 'addparts':
                assetidlist= request.POST['id']
                parentid = int(request.POST['parentid'])

                # Todo 向设备添加配件，1.改配件的父级 2.此配件归档 3.修改配件状态为使用中 4. 刷新对应设备的配件列表
                pps= asset_property.objects.get(id=parentid)

                for assetid in assetidlist.split(",")[0:-1]:
                    print(assetid)
                    ps = asset_property.objects.get(id=int(assetid))

                    ps.parentid = pps
                    # ps.active = False
                    ps.status = 2

                    ps.save()

                data={}

                data['code'] = 0
                data['msg'] = "ok"

                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            if request.POST['act'] == 'delparts':
                assetid = int(request.POST['id'])

                ps = asset_property.objects.get(id=assetid)
                ps.parentid = None
                # ps.active = True
                ps.status = 1
                ps.save()

                data = {}
                data['code'] = 0
                data['msg'] = "ok"

                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
    return render(request, 'asset_parts_form.html', context)

def asset_category_list(request):
    request.encoding = 'utf-8'
    context={}
    context['title']='类型列表'

    username = request.COOKIES.get('usercookie', None)
    if username:
        try:
            signuser = hr_hr.objects.get(session=username)
        except Exception:
            context['userinfo'] = '用户'
            return render(request, 'sign.html', context)
        context['userinfo'] = signuser.name
    else:
        context['userinfo'] = '用户'
        return render(request, 'sign.html', context)

    print(request.method)
    if request.method == "GET":
        if "act" in request.GET:
            if request.GET['act'] == 'sort':
                pass
        else:
            ps = asset_category.objects.filter(active=True).order_by('displayname')
            lcats = []
            for cat in ps:
                ncat = {}
                ncat["id"] = cat.id
                ncat["name"] = cat.name
                ncat["displayname"] = cat.displayname
                ncat["bom"] = cat.bom
                ncat["autosid"] = cat.autosid
                lcats.append(ncat)
            context['context'] = lcats
    return render(request, 'asset_category_list.html', context)

def asset_category_form(request):
    request.encoding = 'utf-8'
    context={}
    context['title']='类型表单'

    username = request.COOKIES.get('usercookie', None)
    if username:
        try:
            signuser = hr_hr.objects.get(session=username)
        except Exception:
            context['userinfo'] = '用户'
            return render(request, 'sign.html', context)
        context['userinfo'] = signuser.name
    else:
        context['userinfo'] = '用户'
        return render(request, 'sign.html', context)

    if request.method == "GET":
        print(request.GET)
        if "act" in request.GET:
            if request.GET['act'] == 'display':
                context['act'] = "display"

                pn = int(request.GET['id'])
                ppn = pn-1
                npn = pn+1
                context['pk'] = pn
                context['ppk'] = ppn
                context['npk'] = npn
                context['spk'] = asset_category.objects.filter(active=True).count()

                cats = asset_category.objects.exclude(id=pn).filter(active=True).order_by('displayname')
                lcats = []
                for cat in cats:
                    ncat = {}
                    ncat["id"] = cat.id
                    ncat["name"] = cat.name
                    ncat["displayname"] = partt(cat.id)
                    lcats.append(ncat)
                context['cats'] = lcats

                try:
                    # 当前设备信息
                    ps = asset_category.objects.get(id=pn)
                except Exception:
                    return redirect('/category_list/')
                else:
                    context['context'] = ps

                assetsub = asset_property.objects.filter(categoryid=ps,active=True).values()
                # print(assetsub)
                context['subnum'] = assetsub.count()
            elif request.GET['act'] == 'create':
                # Todo 新建和复制
                context['act'] = "create"
                context['pk'] = id = int(request.GET['id'])

                cats = asset_category.objects.exclude(id=id).filter(active=True).order_by('parentid', 'name')
                lcats = []
                for cat in cats:
                    ncat = {}
                    ncat["id"] = cat.id
                    ncat["name"] = cat.name
                    ncat["displayname"] = partt(cat.id)
                    lcats.append(ncat)
                context['cats'] = lcats

                if id:
                    pass
                else:
                    pass

    elif request.method == "POST":
        print(request.POST)
        if "act" in request.POST:
            if request.POST['act'] == "active":
                putid = int(request.POST['id'])
                act = asset_category.objects.filter(id=putid).update(active=True)
                return HttpResponse(act)
            if request.POST['act'] == 'unactive':
                putid = int(request.POST['id'])
                act = asset_category.objects.filter(id=putid).update(active=False)
                return HttpResponse(act)
            if request.POST['act'] == "edit":
                act='edit'

                id = int(request.POST['id'])
                catsid = asset_category.objects.get(id=id)

                parentid = int(request.POST['parentid'])
                if parentid:
                    catsid.parentid = asset_category.objects.filter(id=parentid).first()
                else:
                    catsid.parentid = None

                if (request.POST.get('bom', False)):
                    catsid.bom = True
                else:
                    catsid.bom = False

                catsid.name = request.POST['name']
                if catsid.parentid:
                    catsid.displayname = partt(catsid.parentid.id) + ' / ' + catsid.name
                else:
                    catsid.displayname = catsid.name
                catsid.autosid = request.POST['bname']
                catsid.notes = request.POST['notes']

                k = catsid.save()

                ps = asset_category.objects.all()

                for p in ps:
                    if p.parentid:
                        p.displayname = partt(p.parentid.id) + ' / ' + p.name
                    else:
                        p.displayname = p.name
                    p.save()

                data = {}
                if True:
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['data'] = None
                else:
                    data['code'] = 1
                    data['msg'] = "FAIL"
                    data['data'] = None
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")
            if request.POST['act'] == "aplsub":
                id = int(request.POST['id'])
                item = asset_category.objects.get(id=id)
                bomnew = item.bom

                item_u =asset_property.objects.filter(categoryid=item).update(bom=bomnew)
                data = {}
                if item_u:
                    data['code'] = 0
                    data['msg'] = "OK"
                    data['data'] = item_u
                else:
                    data['code'] = 1
                    data['msg'] = "FAIL"
                    data['data'] = None
                data = json.dumps(data)
                return HttpResponse(data, content_type="application/json")

    return render(request, 'asset_category_form.html', context)

def dep_view(request):
    context={}
    context['title']='部门'

    username = request.COOKIES.get('usercookie', None)
    if username:
        try:
            signuser = hr_hr.objects.get(session=username)
        except Exception:
            context['userinfo'] = '用户'
            return render(request, 'sign.html', context)
        context['userinfo'] = signuser.name
    else:
        context['userinfo'] = '用户'
        return render(request, 'sign.html', context)

    ps = hr_department.objects.all().order_by('pid')
    context['context']= ps
    return render(request, 'view_dep_list.html', context)

def hr_view(request):
    context={}
    context['title']='人员'

    username = request.COOKIES.get('usercookie', None)
    if username:
        try:
            signuser = hr_hr.objects.get(session=username)
        except Exception:
            context['userinfo'] = '用户'
            return render(request, 'sign.html', context)
        context['userinfo'] = signuser.name
    else:
        context['userinfo'] = '用户'
        return render(request, 'sign.html', context)

    # d = base_conf.objects.all().first()
    # ps = hr_department.objects.all().order_by('parentid','order')
    # ps = hr_department.objects.filter(parentid=1).order_by('pid', 'order')
    ps = hr_hr.objects.filter(active=True).order_by('name')
    context['context']= ps
    return render(request, 'view_hr_broad.html', context)

def sign_view(request):
    context={}
    context['userinfo'] = '登录'
    context['stat'] = 'CLItMYby44wD8vgM'
    request.encoding = 'utf-8'
    if request.method == "POST":
        if request.POST["user"]:
            seluser = request.POST['user']
            users = hr_hr.objects.filter(userid = seluser)
            if users:
                u = users.first()
                rnd = seluser + str(time.time()) + str(random.randint(10000,20000))
                strsession = hashlib.md5()
                strsession.update(rnd.encode('utf-8'))
                s = strsession.hexdigest()
                u.session = s
                u.save()

                #登录日志
                wlog = base_user_sign_log(signtime = datetime.datetime.fromtimestamp(time.time()),
                                          employeeid = u,
                                          fromip = request.META.get('REMOTE_ADDR','0.0.0.0'),
                                          contl = 'user&passwd')
                wlog.save()

                gourl = redirect('/')
                gourl.set_cookie('usercookie',s,14400)

                return gourl
    elif request.method == "GET":
        if 'code' in request.GET:
            if request.GET['code']:
                code = request.GET['code']

                url = 'https://qyapi.weixin.qq.com/cgi-bin/gettoken'
                v = {}
                v['corpid'] = 'ww74c5af840cdd5cb6'
                v['corpsecret'] = 'HBzQYMHZHw1UwQcqDI8GBsTnTTRJA_ODkgZuo2QuT28'

                r = requests.get(url, params=v)
                t = json.loads(r.text)

                url = 'https://qyapi.weixin.qq.com/cgi-bin/user/getuserinfo'
                v = {}
                v['access_token'] = t['access_token']
                v['code'] = code

                r = requests.get(url, params=v)
                t = json.loads(r.text)

                seluser = t['UserId']
                users = hr_hr.objects.filter(userid=seluser)
                if users:
                    u = users.first()
                    rnd = seluser + str(time.time()) + str(random.randint(10000, 20000))
                    strsession = hashlib.md5()
                    strsession.update(rnd.encode('utf-8'))
                    s = strsession.hexdigest()
                    u.session = s
                    u.save()

                    # 登录日志
                    wlog = base_user_sign_log(signtime=datetime.datetime.fromtimestamp(time.time()),
                                              employeeid=u,
                                              fromip=request.META.get('REMOTE_ADDR', '0.0.0.0'),
                                              contl='Scan WX-code')
                    wlog.save()

                    gourl = redirect('/')
                    gourl.set_cookie('usercookie', s, 14400)
                    return gourl
        else:
            print('code is no')

    return render(request, 'sign.html', context)

def signout(request):
    username = request.COOKIES.get('usercookie', None)
    try:
        signuser = hr_hr.objects.get(session=username)
        signuser.session = '0000000000000000'
        # signuser.expsession = 0
        signuser.save()

        gourl = redirect('/')
        gourl.delete_cookie('username')
        return gourl
    except Exception:
        return redirect('/')
    return redirect('/')

def config(request):
    context={}
    context['title']='设置'

    username = request.COOKIES.get('usercookie', None)
    if username:
        try:
            signuser = hr_hr.objects.get(session=username)
        except Exception:
            context['userinfo'] = '用户'
            return render(request, 'sign.html', context)
        context['userinfo'] = signuser.name
    else:
        context['userinfo'] = '用户'
        return render(request, 'sign.html', context)

    p = base_conf.objects.get(id=1)
    context['context'] = p
    return render(request, 'base_conf.html', context)

def view_pure_list(request):
    context={}
    context['title']='pure list'

    username = request.COOKIES.get('usercookie', None)
    if username:
        try:
            signuser = hr_hr.objects.get(session=username)
        except Exception:
            context['userinfo'] = '用户'
            return render(request, 'sign.html', context)
        context['userinfo'] = signuser.name
    else:
        context['userinfo'] = '用户'
        return render(request, 'sign.html', context)

    ps = pureftp.objects.all().order_by('-id')
    context['context'] = ps
    return render(request, 'asset_pure_list.html', context)

def pure_form(request):
    request.encoding='utf-8'
    context={}
    context['title']='表单'

    username = request.COOKIES.get('usercookie', None)
    if username:
        try:
            signuser = hr_hr.objects.get(session=username)
        except Exception:
            context['userinfo'] = '用户'
            return render(request, 'sign.html', context)
        context['userinfo'] = signuser.name
    else:
        context['userinfo'] = '用户'
        return render(request, 'sign.html', context)

    if int(request.GET['act']):
        p = pureftp.objects.get(id=request.GET['act'])
        context['context'] = p
    else:
        cdate = datetime.date.fromtimestamp(time.time())
        ldate = datetime.date.fromtimestamp(time.time())

        p = {'status': 'true',
             'ipaccess': '0.0.0.0',
             'uid': 1001,
             'gid': 1001,
             'ulbandwidth': 0,
             'dlbandwidth': 0,
             'quotasize': 0,
             'quotafiles': 0,
             'createdate': cdate,
             'lastedate': ldate}

        context['context'] = p

    context['act'] = request.GET['act']
    return render(request, 'pure_form.html', context)

def pure_add(request):
    request.encoding = 'utf-8'
    if request.method == "POST":
        if int(request.POST["acte"]):
            cid = int((request.POST["acte"]))

            if 'status' in request.POST:
                cstatus = 1
            else:
                cstatus = 0
            cuser = str(request.POST['user'])
            cpassword = str(request.POST['password'])
            cipaccess = str(request.POST['ipaccess'])
            cdir = str(request.POST['dir'])
            cuid = str(request.POST['uid'])
            cgid = str(request.POST['gid'])
            culbandwidth = int(request.POST['ulbandwidth'])
            cdlbandwidth = int(request.POST['dlbandwidth'])
            cquotasize = int(request.POST['quotasize'])
            cquotafiles = int(request.POST['quotafiles'])
            ccomment = str(request.POST['comment'])

            pureftp.objects.filter(id=cid).update(status = cstatus,
                                                  user = cuser,
                                                  password = cpassword,
                                                  ipaccess = cipaccess,
                                                  dir = cdir,
                                                  uid = cuid,
                                                  gid = cgid,
                                                  ulbandwidth = culbandwidth,
                                                  dlbandwidth = cdlbandwidth,
                                                  quotafiles = cquotafiles,
                                                  quotasize = cquotasize,
                                                  comment = ccomment)
            return redirect('/pure_list/')
        else:
            addnew = pureftp(
                status = 1,
                user = str(request.POST['user']),
                password = str(request.POST['password']),
                ipaccess = str(request.POST['ipaccess']),
                dir = str(request.POST['dir']),
                uid = str(request.POST['uid']),
                gid = str(request.POST['gid']),
                ulbandwidth = int(request.POST['ulbandwidth']),
                dlbandwidth = int(request.POST['dlbandwidth']),
                quotasize = int(request.POST['quotasize']),
                quotafiles = int(request.POST['quotafiles']),
                createdate = str(request.POST['createdate']),
                lastedate = str(request.POST['lastedate']),
                comment = str(request.POST['comment']))
            addnew.save()
            return redirect('/pure_list/')

def pure_del(request):
    request.encoding='utf-8'
    context={}
    context['title']='pure form'

    if "act" in request.GET:
        delid = int(request.GET['id'])
        delitem = pureftp.objects.get(id=delid)
        delitem.delete()

    return redirect('/pure_list/')

def getToken(request):
    if "act" in request.GET:
        if request.GET['act'] == 'gettoken':
            # list = pureftp.objects.all()
            # d = base_conf.objects.all().first()
            # d=base_conf.objects.filter(id=1)

            # d=base_conf.objects.get(id=1)
            d = base_conf.objects.get(pk=1)

            # pureftp.objects.order_by("id")

            #https://qyapi.weixin.qq.com/cgi-bin/gettoken?corpid=ID&corpsecret=SECRECT
            # corpid=ww74c5af840cdd5cb6
            # corpsecret=uUJf-eFplyKlAWf2Cc9T8Guea4K1zpEiZXwXpCsHTQs

            url = 'https://qyapi.weixin.qq.com/cgi-bin/gettoken'
            v = {}
            v['corpid'] = d.corpid
            v['corpsecret'] = d.corpsecret

            r = requests.get(url, params=v)
            t = json.loads(r.text)

            if t['errcode'] == 0:
                d.token = t['access_token']

                tt = time.time()
                d.expirestime = datetime.datetime.fromtimestamp(tt + t['expires_in'])

                d.save()
                response = d.token
            else:
                response = "Error"

    return HttpResponse(response)

def readdepartment(request):
    p =  base_conf.objects.all().first()

    # https://qyapi.weixin.qq.com/cgi-bin/department/list?access_token=ACCESS_TOKEN&id=ID
    url = 'https://qyapi.weixin.qq.com/cgi-bin/department/list'
    v = {}
    v['access_token'] = p.token
    v['id'] = p.corpsecret

    try:
        r = requests.get(url, params=v)
    except Exception:
        print('get url is error')
    else:
        t = json.loads(r.text)
        if t['errcode'] == 0:
            d ={}
            for d in t['department']:
                dt = hr_department(pid=d['id'],
                                   name=d['name'],
                                   parentid=d['parentid'],
                                   order=d['order'])
                dt.save()
        response = t
    return redirect('/dep/')

def gethr(request):
    p =  base_conf.objects.all().first()

    # https://qyapi.weixin.qq.com/cgi-bin/user/list?access_token=ACCESS_TOKEN&department_id=DEPARTMENT_ID&fetch_child=FETCH_CHILD
    url = 'https://qyapi.weixin.qq.com/cgi-bin/user/list'
    # https://qyapi.weixin.qq.com/cgi-bin/user/simplelist?access_token=ACCESS_TOKEN&department_id=DEPARTMENT_ID&fetch_child=FETCH_CHILD
    # url = 'https://qyapi.weixin.qq.com/cgi-bin/user/simplelist'
    v = {}
    v['access_token'] = p.token
    v['department_id'] = p.department_id =1
    v['fetch_child'] = 1

    r = requests.get(url, params=v)
    t = json.loads(r.text)

    if t['errcode'] == 0:
        d = {}

        for d in t['userlist']:
            print('-----------------')
            try:
                chkem=None
                chkem = hr_hr.objects.filter(userid=d['userid'])
            except Exception:
                print('chkem error')
                pass
            if chkem:
                #记录已存在
                pass
            else:
                #-----------------
                u_userid = d['userid']
                u_name = d['name']
                u_position = d['position']
                u_mobile = d['mobile']
                u_gender = d['gender']
                u_email = d['email']
                u_avatar = d['avatar']
                u_status = d['status']
                u_enable = d['enable']
                u_isleader = d['isleader']
                u_extattr = d['extattr']
                u_hide_mobile = d['hide_mobile']
                u_english_name = d['english_name']
                u_telephone = d['telephone']
                u_order = d['order']
                # u_external_profile=d['external_profile']
                u_qr_code = d['qr_code']
                #-----------------

                dt = hr_hr(userid=d['userid'],
                           name=u_name,
                           position=u_position,
                           mobile=u_mobile,
                           gender=u_gender,
                           email=u_email,
                           avatar=u_avatar,
                           status=u_status,
                           enable=u_enable,
                           isleader=u_isleader,
                           extattr=u_extattr,
                           hide_mobile=u_hide_mobile,
                           english_name=u_english_name,
                           telephone=u_telephone,
                           order=u_order,
                           # external_profile=u_external_profile,
                            qr_code=u_qr_code)
                dt.save()
                print(u_name)

                #关联员工与部门
                nemp = hr_hr.objects.get(userid=d['userid'])
                ds = d['department']

                for de in ds:
                    des = int(de)
                    try:
                        dep = hr_department.objects.filter(pid=des).first()
                    except Exception:
                        pass
                    else:
                        try:
                            rep = employee_department.objects.filter(employeeid=nemp.userid,departmentid=des)
                        except Exception:
                            pass
                        else:
                            if len(rep):
                                pass
                            else:
                                seds = employee_department(employeeid=nemp,departmentid=dep)
                                seds.save()

    return redirect('/hr/')

def uploadfile(request):
    print(request.POST)
    if request.method == "POST":
        f = request.FILES["file"]
        filePath = os.path.join(settings.MDEIA_ROOT, f.name)
        print(filePath)
        with open(filePath,'wb') as fp:
            for info in f.chunks():
                fp.write(info)

        wb = load_workbook(filePath)
        e = wb.sheetnames
        print(e)
        sheet = wb["Sheet2"]
        for row in sheet.rows:
            for cell in row:
                print(cell.value)

        return HttpResponse('ok')
    else:
        return HttpResponse('no')

def sendmsg(request):
    v = {}
    v['access_token'] = 'PP-YvltOWYOlkz28puKRqlCIA8pkrGy2X-qMapexdjz-2-CmRA8eteyZ1g2rUzu5IWS3lnmoIHX1nR5EZTEKJb2RR6WroAR-KHvl0Zl3Al886Ny-pySOqkP8obzTQlw1ipMVRTZ1wAGpXW3vt3mLCTkCCo2unZ5f_oPjZzRmU100QOTEmWyB3a0Zi50uwVk5BNzk9YP8PSK-wlNdWFkheQ'
    txt = {"touser" : "SunJun",
               "msgtype" : "text",
               "agentid" : 1000013,
               "text" : {
                   "content" : "测试消息\n文本内容<a href=\"http://work.weixin.qq.com\">网页链接</a>，请勿回复。"},
               "safe":0
               }
    # jsontxt = json.dumps(context)
    print(txt)

    urls = 'https://qyapi.weixin.qq.com/cgi-bin/message/send'
    r = requests.post(urls,params=v,json=txt)

    return HttpResponse(r.text)

def wxdoor(request):
    # url = 'https://qyapi.weixin.qq.com/cgi-bin/gettoken'
    # v = {}
    # v['corpid'] = 'ww74c5af840cdd5cb6'
    # v['corpsecret'] = 'HBzQYMHZHw1UwQcqDI8GBsTnTTRJA_ODkgZuo2QuT28'
    #
    # r = requests.get(url, params=v)
    # t = json.loads(r.text)
    # print(t['access_token'])
    #
    # context = t['access_token']

    # print(request.text)
    # v = {}
    # url = 'https://open.weixin.qq.com/connect/oauth2/authorize'
    # v['appid'] = 'ww74c5af840cdd5cb6'
    # v['redirect_uri'] = '172.18.0.231:8000'
    # v['response_type'] = 'code'
    # v['scope'] = 'snsapi_base'
    # v['agentid'] = '1000013'
    # v['state'] =  ''
    # r = requests.get(url, params=v)

    # print(request.method)
    # print(request.GET)
    # print(request.COOKIES)
    # g = get hr_hr.objects

    url = 'https://open.weixin.qq.com/connect/oauth2/authorize?appid=ww74c5af840cdd5cb6&redirect_uri=http://p.jtanimation.com:8000/wxcode&response_type=code&scope=snsapi_base&agentid=1000013&state=STATE#wechat_redirect'
    print(url)
    return redirect(url)

def wxcode(request):
    if request.method  == "GET":
        if 'code' in request.GET:
            code = request.GET["code"]

            url = 'https://qyapi.weixin.qq.com/cgi-bin/gettoken'
            v = {}
            v['corpid'] = 'ww74c5af840cdd5cb6'
            v['corpsecret'] = 'HBzQYMHZHw1UwQcqDI8GBsTnTTRJA_ODkgZuo2QuT28'

            r = requests.get(url, params=v)
            t = json.loads(r.text)

            url = 'https://qyapi.weixin.qq.com/cgi-bin/user/getuserinfo'
            v = {}
            v['access_token'] = t['access_token']
            v['code'] = code

            print(url)

            r = requests.get(url, params=v)
            t = json.loads(r.text)

            print(t)

            seluser = t['UserId']
            users = hr_hr.objects.filter(userid=seluser)
            if users:
                u = users.first()
                rnd = seluser + str(time.time()) + str(random.randint(10000, 20000))
                strsession = hashlib.md5()
                strsession.update(rnd.encode('utf-8'))
                s = strsession.hexdigest()
                u.session = s
                u.save()

                # 企业微信通过后跳转的首页
                gourl = redirect('/asset/')
                gourl.set_cookie('usercookie', s, 14400)
                return gourl

    return HttpResponse('wxcode')

def wxtext(request):
    return HttpResponse('nHQMbL9kxBaRLdjd')

def property_upload(request):
    print(request.POST)
    if request.method == "POST":
        f = request.FILES["file"]
        # 生成新的文件名
        fext = os.path.splitext(f.name)[1]
        assetsn = request.POST["sn"]
        assetid = request.POST["id"]
        assetnum = asset_attachment.objects.filter(property_id=assetid).count()
        assetnum += 1
        nname = "%s-%02d%s" % (assetsn, assetnum, fext)

        filePath = os.path.join(settings.MDEIA_ROOT, 'property\img', nname)
        with open(filePath,'wb') as fp:
            for info in f.chunks():
                fp.write(info)

        pid = int(request.POST['id'])
        try:
            propertyid = asset_property.objects.get(id=pid)
        except Exception:
            return HttpResponse('data no')

        asset_attachment.objects.filter(property=pid).update(final=False)

        # 网站中的相对文件
        pth='/static/upfile/property/img/' + nname

        # 预览图的路径及相对路径
        thumbnail= os.path.join(settings.MDEIA_ROOT, 'property\img\\thumbnail\\64', nname)
        nameth = '/static/upfile/property/img/thumbnail/64/' + nname

        # 转小文件
        im = Image.open(filePath)
        thu = im.resize((64,64))
        thu.save(thumbnail)

        k = asset_attachment(property=propertyid,
                             name=nname,
                             filepath=pth,
                             thumbnail=nameth,
                             oldname=f.name,
                             version=0,
                             final=True,
                             category='0')
        k.save()
        data= {}
        data['id'] = k.id
        data['filepath'] = k.filepath

        data = json.dumps(data)
        return HttpResponse(data, content_type="application/json")
    else:
        return HttpResponse('no')

def asset_property_list_output_cvs(request):
    response = HttpResponse(content_type='text/csv')
    response.write(codecs.BOM_UTF8)
    response['Content-Disposition'] = 'attachment; filename="somefilename.csv"'

    ps = asset_property.objects.filter(bom=False, active=True).filter(Q(asset_attachment__final=True)
                                                                      | Q(asset_attachment__final=None)).values_list('status',
                                                                                                                'sid',
                                                                                                                'name',
                                                                                                                'specifications',
                                                                                                                'warranty',
                                                                                                                'user__name',
                                                                                                                'position',
                                                                                                                'user__employee_department__departmentid__name',
                                                                                                                'sn').order_by('name', 'sid')
    writer = csv.writer(response)

    th = ['状态', '编号', '名称', '规格', '维保到期', '用户', '部门', '位置', '出厂编号']
    writer.writerow(th)
    for i in ps:
        li = list(i)
        li[0] = status(li[0],0)
        writer.writerow(li)
    # writer.writerow(['First row', 'Foo', 'Bar', 'Baz',2,"3"])
    # writer.writerow(['Second row', 'A', 'B', 'C', '"Testing"', "Here's a quote",1,"5"])

    # csv_data = (
    #     ('First row', 'Foo', 'Bar', 'Baz'),
    #     ('Second row', 'A', 'B', 'C', '"Testing"', "Here's a quote"),
    # )
    #
    # t = loader.get_template('my_template_name.txt')
    # c = Context({
    #     'data': csv_data,
    # })
    # response.write(t.render(c))

    # print(response.content)
    return response

#功能测试路由
def importdata(request):
    url = 'https://www.qqxiuzi.cn/zh/pinyin/'
    # r = requests.get(url, params=v)
    r = requests.get(url)
    r.encoding = "utf-8"

    print(r.text)

    for key, value in r.cookies.items():
        print(key, '==', value)
    # tk = re.match(r"thoken",r.text,flags=0)
    # print(tk)

    url = 'https://www.qqxiuzi.cn/zh/pinyin/show.php'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:63.0) Gecko/20100101 Firefox/63.0',
    }

    v = {}
    v['t'] = '胜多负少'
    v['d'] = 3
    v['s'] = 'null'
    v['k'] = 1
    v['b'] = 'null'
    v['h'] = 'null'
    v['u'] = 'null'
    v['v'] = 'null'
    v['y'] = 'null'
    v['z'] = 'null'
    v['f'] = 'null'
    v['token'] = '97fb41d9c78f12787d6e66405a21576b'

    r1 = requests.post(url,cookies=r.cookies, headers=headers, params=v)
    r1.encoding = "utf-8"
    print(r1.url)
    print(r1.cookies)
    print(r1.text)
    print(r1.status_code)


    # url = 'http://www.atool.org/include/pinyin.inc.php'
    # # url = 'http://httpbin.org/post'
    # headers = {
    #     'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:63.0) Gecko/20100101 Firefox/63.0'
    # }
    #
    # v = {}
    # v['w'] = '胜多负少'
    # v['d'] = '-'
    # v['sd'] = 'no'
    # v['v'] = 9
    #
    # r = requests.post(url, params=v)
    # r.encoding = "utf-8"
    # print(r.url)
    # print(r.text)
    # print(r.status_code)

    # wb = load_workbook("kkk.xlsx")

    # # 导入设备
    # sheet = wb["Sheet1"]
    # print(sheet["G1"].value)
    #
    # p = list()
    # x = 0
    # for i in sheet["A"]:
    #     x += 1
    #     if x != 1:
    #         u = sheet["G"+str(x)].value
    #         m = sheet["J"+str(x)].value
    #         w = sheet["K"+str(x)].value
    #
    #         cat = sheet["E" + str(x)].value
    #         catn = asset_category.objects.filter(name=cat).first()
    #
    #         c = sheet["S" + str(x)].value
    #         cou = hr_hr.objects.filter(name=c).first()
    #
    #         p.append(asset_property(
    #             sid=sheet["B"+str(x)].value,
    #             name=sheet["C"+str(x)].value,
    #             specifications=sheet["F"+str(x)].value,
    #             model=sheet["U"+str(x)].value,
    #             # categoryid=sheet["E"+str(x)].value,
    #             categoryid=catn,
    #             purchase=u.date(),
    #             price=sheet["O"+str(x)].value,
    #             manufacture=m.date(),
    #             warranty=w.date(),
    #             sn=sheet["I"+str(x)].value,
    #             status=sheet["Q" + str(x)].value,
    #             position=sheet["R" + str(x)].value,
    #             # user=sheet["Q"+str(x)].value,
    #             user=cou,
    #             # partlist=1,
    #             # notes=sheet["V" + str(x)].value,
    #         ))
    # # print(p)
    # asset_property.objects.bulk_create(p)



    # # 导入设备使用人
    # x = 0
    # m = 0
    # n = 0
    # w = 0
    # for i in sheet["A"]:
    #     x += 1
    #     if x != 1:
    #         s = sheet["B" + str(x)].value
    #         c =sheet["S"+str(x)].value
    #
    #         cou = hr_hr.objects.filter(name=c).first()
    #         if c:
    #             if cou:
    #                 w += 1
    #                 # asset_property.objects.filter(sid=s).update(user=cou)
    #             else:
    #                 m += 1
    #         else:
    #             n += 1
    # print("空=%s   无=%s   写=%s"% (n,m,w))

    # # 导入类型
    # sheet1 = wb["Sheet2"]
    # n = 0
    # for i in sheet1['A']:
    #     n += 1
    #     if n != 1:
    #         namen = sheet1["B" + str(n)].value
    #         upd = sheet1["G" + str(n)].value
    #         upd = int(upd)+1
    #         up = sheet1["B" + str(upd)].value
    #         print("%s - %s" % (namen,up))
    #
    #         ni = asset_category.objects.filter(name=namen)
    #         if ni:
    #             puid = asset_category.objects.filter(name=up).first()
    #             asset_category.objects.filter(name=namen).update(parentid=puid)
    #         else:
    #             #add
    #             nn = asset_category(name=namen)
    #             nn.save()
    #             puid = asset_category.objects.filter(name=up).first()
    #             asset_category.objects.filter(name=namen).update(parentid=puid)

    # # 更新在用
    # it = asset_property.objects.filter(user__isnull=False).update(status=2)

    # # 导入类型
    # sheet1 = wb["Sheet2"]
    # n = 0
    # for i in sheet1['A']:
    #     n += 1
    #     if n != 1:
    #         ids = sheet1["B" + str(n)].value
    #         cat = sheet1["E" + str(n)].value
    #
    #         catn = asset_category.objects.filter(name=cat).first()
    #
    #         asset_property.objects.filter(sid=ids).update(categoryid=catn)
    #         print("%s - %s" % (ids,cat))

    # its = asset_property.objects.values_list('categoryid').annotate(Count('id'))
    # print(its.query)
    #
    # for it in its:.
    #     print(it)
    # print(type(it))

    # if request.method == 'POST':
    #     print(request.POST)

    # # 导入配件
    # sheet = wb["Sheet3"]
    # print(sheet["G2"].value)
    #
    # p = list()
    # x = 0
    # for i in sheet["A"]:
    #     x += 1
    #     if x != 1:
    #
    #         u = sheet["D"+str(x)].value
    #         m = sheet["D"+str(x)].value
    #         w = sheet["L"+str(x)].value
    #
    #         # print(type(u))
    #         print(sheet["A"+str(x)].value)
    #         # print(type(datetime.datetime.strptime(u, '%Y/%m/%d %H:%M:%S')))
    #
    #         cat = sheet["H" + str(x)].value
    #         catn = asset_category.objects.filter(name=cat).first()
    #
    #         print(catn)
    #
    #         # c = sheet["S" + str(x)].value
    #         # cou = hr_hr.objects.filter(name=c).first()
    #
    #         # p = list()
    #         # p.append(asset_property(
    #         #     sid=sheet["AA"+str(x)].value,
    #         #     name=sheet["E"+str(x)].value,
    #         #     specifications=sheet["F"+str(x)].value,
    #         #     # model=sheet["U"+str(x)].value,
    #         #     # categoryid=sheet["E"+str(x)].value,
    #         #     categoryid=catn,
    #         #     purchase=u.date(),
    #         #     price=sheet["Q"+str(x)].value,
    #         #     manufacture=m.date(),
    #         #     warranty=w.date(),
    #         #     sn=sheet["Z"+str(x)].value,
    #         #     status=sheet["C" + str(x)].value,
    #         #     position="18F技术部",
    #         #     # user=sheet["Q"+str(x)].value,
    #         #     # user=cou,
    #         #     # partlist=1,
    #         #     nots=sheet["K" + str(x)].value,
    #         # ))
    #         # asset_property.objects.bulk_create(p)
    # # asset_property.objects.bulk_create(p)

    # # 导入价格
    # sheet = wb["Sheet5"]
    #
    # x = 0
    # n = 0
    # sumi = 0
    # for i in sheet["A"]:
    #     x += 1
    #     if x != 1:
    #         s = sheet["F"+str(x)].value
    #         p = 0
    #         if s != None:
    #             p = sheet["G"+str(x)].value
    #             d = sheet["I"+str(x)].value
    #             if p != 0:
    #                 ps = asset_property.objects.filter(sn__icontains=s).first()
    #                 if ps:
    #                     # print(ps.sn, ps.price,ps.purchase, s, p, d)
    #                     # ps.price = p
    #                     # n += 1
    #                     # # ps.save()
    #                     # print(n)
    #                     pass
    #                 else:
    #                     n += 1
    #                     sumi = sumi + p
    #                     print(s, p, d)
    #                     pass
    # print(n,sumi)

    response = {"status":"ok"}
    return HttpResponse(response)

def search(request):
    context={}
    context['title']='pure list'

    # username = request.COOKIES.get('usercookie', None)
    # if username:
    #     try:
    #         signuser = hr_hr.objects.get(session=username)
    #     except Exception:
    #         context['userinfo'] = '用户'
    #         return render(request, 'sign.html', context)
    #     context['userinfo'] = signuser.name
    # else:
    #     context['userinfo'] = '用户'
    #     return render(request, 'sign.html', context)
    # print(signuser.id)

    # depid = 8
    # hre = hr_hr.objects.get(id=depid)
    # dpe = employee_department.objects.filter(employeeid__gender=1)
    # print(dpe.query)
    # print('--------------------------')
    # # print(dpe)
    # for i in dpe:
    #     print(i)
    #
    # print('=================================')
    # hre = employee_department.objects.all().values('employeeid__name','departmentid__name')
    # print(hre.query)
    # print('--------------------------')
    # print(hre)
    #
    # print('--------------------------')
    # hee = employee_department.objects.all()
    # for i in hee:
    #     print(i.employeeid.name)

    # em = hr_department.objects.get(pid=8)
    # print(em.employee_department_set.all())

    # h = hr_department.objects.get(pid=9)
    # lks = h.employee_department_set.all()
    # for p in lks:
    #     print(p.employeeid.name)

    # pr = asset_property.objects.get(pk=1)
    # pts = pr.asset_parts_set.all()
    # for pt in pts:
    #     print(pt.name)
    #
    # pr1 = asset_property.objects.filter(pk=1)
    # print(type(pr1), pr1.query)
    #
    # pr = asset_property.objects.get(pk=1)
    # print(type(pr))
    #
    # print(pr.name,
    #       pr.user.name,
    #       pr.user.employee_department_set.all().values('departmentid__name'))

    # q = asset_category.objects.all().values('name')
    # print(q)
    # for iq in q:
    #     print(iq)

    return render(request, 'search.html',context)
    # return render(request, 'view_hr_list.html',context)

def sub(request):
    # context={}
    # context['title']='sub list'
    # ps = hr_hr.objects.filter(active=True).order_by('name')
    # kk = hr_hr.objects.first()
    # # kk.employee_department_set.all()
    # print(kk)
    # context['context'] = ps
    # return render(request, 'hr_list_sub.html', context)

    ps = asset_category.objects.all().values_list()
    # print(ps)

    # for item in ps:
    #     # pass
    #     print(item)
        # print(item[0])

    s = list(ps)
    # print(s)
    # s.sort(key=lambda x:x[2])
    # s.sort()
    # print(type(s))
    # print(s)
    c =[]
    for item in s:
        k = list(item)
        k.append(partt(item[0]))
        # print(type(k))
        # print(k)
        c.append(k)
    c.sort(key=lambda x: x[7])
    for e in c:
        print(e[7])

    response = {"status":"ok"}
    return HttpResponse(response)